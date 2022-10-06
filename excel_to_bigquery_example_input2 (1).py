import pandas as pd
import re
import logging
import glob
from datetime import datetime
from time import perf_counter
import calendar
import os
import numpy as np
from datetime import date
from google.cloud import bigquery
from google.cloud import storage
import argparse
from openpyxl import load_workbook
import io

ts_format = '%(asctime)s - %(name)s - %(threadName)s - %(levelname)s - %(message)s'
logging.basicConfig(level=logging.INFO, format=ts_format)
logger = logging.getLogger(__name__)

user_name = os.environ['LOGNAME']
ts = datetime.utcnow().strftime("%Y%m%d%H%M%S")
now_ts = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
now_dt = datetime.utcnow().strftime("%Y-%m-%d")

def _parse_args():
    parser = argparse.ArgumentParser()
    # parser.add_argument("--full_gcs_path_pattern", required=True, help="example: gs://landing_supplychain-dlf-dev-72acad/sourcing/landing/LGCP_SPEND_FY")
    # parser.add_argument("--bq_table_id", required=True, help="project.schema.table for loading file(s) to")
    # parser.add_argument("--bq_field_datatype", required=False, help="dictionary for field type override")
    # parser.add_argument("--bq_table_filter", required=False, help="")
    args = parser.parse_args()
    return args


def scrub_the_workbook_sheets(full_gcs_path_pattern, sheet_map):
    """
    workbook sheet to sqlite table either specific sheet or all sheets
        also capture workbook metadata
    """

    root_gcs_bucket = full_gcs_path_pattern.replace('gs://','').split('/')[0]                            #landing_supplychain-dlf-dev-72acad
    bucket_subdirectory = full_gcs_path_pattern.replace('gs://','').replace(root_gcs_bucket,'')[1:]          #sourcing/landing/LGCP_SPEND_FY
    print('root_gcs_bucket',root_gcs_bucket)
    print('bucket_subdirectory',bucket_subdirectory)

    storage_client = storage.Client()
    bucket = storage_client.bucket(root_gcs_bucket)
    bq_client = bigquery.Client()
    # bq_job_config = bigquery.LoadJobConfig(write_disposition="WRITE_APPEND",autodetect=False) 
    bq_job_config = bigquery.LoadJobConfig(write_disposition="WRITE_TRUNCATE",autodetect=False) 

    file_list_all = get_filenames(storage_client,bucket)
    file_list = []  
    for file in file_list_all:
        if bucket_subdirectory in file['name']:
            file_list.append(''.join([f'gs://{root_gcs_bucket}/', file['name']]))
    print(file_list)
    if len(file_list)>0:
        current_gcs_project = root_gcs_bucket.replace('user_inputs_','')
        current_subdirectory = bucket_subdirectory.replace(bucket_subdirectory.split('/')[-1],'').replace('/','_')
        # print('current_subdirectory',current_subdirectory)
        sheets_processed = 0
        for file in file_list:
            print('file:',file)
            full_workbook_path = file
            time_beg = perf_counter()
            print('full_workbook_path:',full_workbook_path)            
            try:
                blob = bucket.blob(bucket_subdirectory)
                buffer = io.BytesIO()
                blob.download_to_file(buffer)
                wb = load_workbook(buffer,data_only=True,keep_vba=True) #keep_vba=True for .xlsm
            except Exception as e:
                print('Exception',e)
                print('cant use openpyxl. error occurred')

            time_end = perf_counter()
            print('loaded wb')

            try:
                wb_last_modified_by = wb.properties.lastModifiedBy
                wb_last_modified_date = wb.properties.modified
                print('retrieved modified metadata')
                sheets = wb.sheetnames
                method = 'openpyxl'
            except:
                wb_last_modified_by = ''
                wb_last_modified_date = ''
                print('taken from read_excel... non metadata')
                sheets = wb.keys()
                method = 'pandas'

            for each in sheet_map:
                (tab, r, c) = each
                print('sheets in play:', sheets)
                # sheet_array = []
                for sheet in sheets:
                    if sheet == tab or tab == '':
                        print('starting with:', sheet, 'tab:', tab)
                        dsheet = wb[sheet]
                        if method == 'openpyxl':
                            first_row = dsheet.min_row
                            first_col = dsheet.min_column
                            # last_col =dsheet.max_column
                            # last_row = dsheet.max_row

                            if r == 1 and c == 1:
                                r = first_row
                                c = first_col
                            else:
                                if r != 1:
                                    r = r - 1 #adjust for index at 0
                                if c != 1:
                                    c = c - 1 #adjust for index at 0
                            
                            df = pd.DataFrame(dsheet.values)
                            # df = pd.DataFrame(dsheet.values).astype('string')
                            for col in df.columns:
                                df[col] = df[col].astype('string')

                            new_header = df.iloc[r - 1]  # grab the first row for the header
                            df = df[r:]  # take the data less the header row
                            df.columns = new_header  # set the header row as the df header
                        else:  # this may be used if openpyxl is not available for this workbook, otherwise skip
                            print('simple review')
                            print('dsheet.head(5)', dsheet.head(5))
                            print('r', r)

                            print('len dsheet index:', len(dsheet.index))
                            print('dsheet shape rows:', dsheet.shape[0])
                            print('dsheet shape cols:', dsheet.shape[1])
                            # print('dsheet df rows count   :',dsheet[dsheet.rows[0]].count())
                            print('dsheet df columns count:', dsheet[dsheet.columns[0]].count())
                            # dsheet = dsheet.dropna(how='all')
                            # df = dsheet.reset_index(drop=True)
                            df = dsheet.dropna(how='all')
                            print('len df index:', len(df.index))
                            print('df shape rows:', df.shape[0])
                            print('df shape cols:', df.shape[1])
                            # print('df df rows count   :',df[df.rows[0]].count())
                            print('df df columns count:', df[df.columns[0]].count())

                            if sheet == 'anotherSheet Name':
                                r = 0
                                new_header = df.iloc[r]  # grab the first row for the header
                                df = df[r + 1:]  # take the data less the header row
                                df.columns = new_header  # set the header row as the df header
                            if sheet == 'MySheetName':
                                r = 1
                                new_header = df.iloc[r]  # grab the first row for the header
                                df = df[r + 1:]  # take the data less the header row
                                df.columns = new_header  # set the header row as the df header

                        print('df.head(5)', df.head(5))
                        (df, header_scrubbed) = scrub_column_headers_on_df(df)
                        
                        if r'\\' in full_workbook_path:
                            scrubbed_file_name = str(full_workbook_path).split('\\')[-1]
                        else:
                            scrubbed_file_name = str(full_workbook_path).split('/')[-1]


                        scrubbed_file_name = scrubbed_file_name.replace(" ", "_").replace("-", "_").replace(".", "_").replace("___", "_").replace("__", "_").lower() #.replace(".xlsx","").replace(".xlsm","").replace(".xls", "")
                        scrubbed_file_name = re.sub("[^a-zA-Z_/0-9]+", "", scrubbed_file_name)
                        scrubbed_sheet_name = sheet.replace(" ", "_").replace("-", "_").replace("___", "_").replace("__", "_").lower()
                        scrubbed_sheet_name = re.sub("[^a-zA-Z_/0-9]+", "", scrubbed_sheet_name)

                        bq_table_id = current_gcs_project + '.input.' + current_subdirectory + scrubbed_file_name + '_' + scrubbed_sheet_name

                        df['ss_wb_last_modified_by'] = wb_last_modified_by
                        df['ss_wb_last_modified_date'] = wb_last_modified_date
                        # df['ss_header_set'] = header_scrubbed
                        # df.columns=col_names
                        df['ss_source_reference'] = str(full_workbook_path)
                        df['ss_source_reference_item'] = sheet
                        df['ss_source_reference_subitem'] = '(' + str(r) + ',' + str(c) + ')'
                        df['ss_destination_reference'] = bq_table_id
                        df['ss_load_ts'] = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
                        df['ss_load_dt'] = now_dt
                        df['ss_load_username'] = user_name
                        df.index = df.index + 1  # move index to use index as row counter starting on row 1 (instead of index default of 0)
                        df.index.name = 'ss_source_reference_row'
                        bq_job = bq_client.load_table_from_dataframe(df,bq_table_id,job_config=bq_job_config) 
                        sheets_processed = sheets_processed + 1
                        bq_job.result()
                        bq_table = bq_client.get_table(bq_table_id)
                        print('bq_table_id:',bq_table_id)
                        time_end = perf_counter()
                        process_time_in_seconds = round(time_end - time_beg, 4)
                        print('process_time:',process_time_in_seconds)
                        print("Loaded {} rows and {} columns to {} from {} taking {} seconds.".format(bq_table.num_rows, len(bq_table.schema), bq_table_id, file,process_time_in_seconds)       ) 
                        # print('df',df)
                        r = 1
                        c = 1
            try:
                wb.close()
            except:
                print('pandas doesnt need to close workbook')


def get_filenames(storage_client,bucket):
    return [{'name': blob.name,
             'type': blob.content_type} for blob in list(storage_client.list_blobs(bucket))] 

def coalesce(*arg): return next((a for a in arg if a is not None), None)

def main():
    
    parsed_args = _parse_args()
    # results = scrub_the_workbook_sheets('gs://input_sc-gbl-impact-dna-dev-4b4e15/user_input/template_data.xlsx',[('numbers',1,1),('',1,1)])
    results = scrub_the_workbook_sheets('gs://user_inputs_sc-gbl-impact-dna-dev-4b4e15/load/template_data.xlsx',[('numbers',1,1),('',1,1)])
    
    logger.info("execution completed")

def scrub_column_headers_on_df(df):
    fnames = []
    # df.columns = df.iloc[0, :]
    # df.columns = df.iloc[0]
    header_array = df.columns
    header_scrubbed = ''
    i = 1
    for cname in header_array:
        cname = str(cname).strip().lower()
        if cname[:1].isdigit():
            cname = 'n' + cname
        cname = cname.replace("\r", "")
        cname = cname.replace("\n", "")
        cname = cname.replace('""', "")
        cname = cname.replace(' ', "_")
        cname = cname.replace(";", "")
        cname = cname.replace("=", "")
        cname = cname.replace("$", "amt")
        cname = cname.replace(",", "_")
        cname = cname.replace("/", "_")
        cname = cname.replace("(", "")
        cname = cname.replace(")", "")
        cname = cname.replace("’", "")
        cname = cname.replace(":", "")
        cname = cname.replace("-", "_")
        cname = cname.replace("_-_", "_")
        cname = re.sub('-', '_', cname)
        cname = cname.replace("–", "_")
        cname = cname.replace("'", "")
        cname = cname.replace("?", "")
        cname = cname.replace("ó", "o")
        cname = cname.replace("Ó", "O")
        cname = cname.replace("°", "")
        cname = cname.replace("[", "")
        cname = cname.replace("]", "")
        cname = cname.replace("|", "")
        cname = cname.replace(" ", "_")
        cname = cname.replace("<", "")
        cname = cname.replace(">", "")
        cname = cname.replace("#", "nbr")
        cname = cname.replace("*", "")
        cname = cname.replace("+", "plus")
        cname = cname.replace("%", "pcnt")
        cname = cname.replace(".", "_")
        cname = cname.replace("&", "_and_")

        cname = re.sub("[^a-zA-Z_0-9]+", "", cname)  # only allow ASCII A-Za-z0-9 and underscore
        cname = cname.replace("_____", "_")
        cname = cname.replace("____", "_")
        cname = cname.replace("___", "_")
        cname = cname.replace("__", "_")

        if cname.strip("_") == "":
            cname = 'c' + str(i)
        if cname[-1] == '_':
            cname = cname[:-1]  # strip last underscore if exists
        iter_count = 1
        col_scrub = ""
        for fname in fnames:
            if cname.lower() == fname.lower() or col_scrub.lower() == fname.lower():
                iter_count += 1
                col_scrub = cname + '_c' + str(iter_count)
            else:
                continue

        if i == 1:  # first column
            if col_scrub == "":
                fnames.append(cname)
                header_scrubbed = cname
            else:
                fnames.append(col_scrub)
                header_scrubbed = header_scrubbed + ',' + col_scrub
        else:  # all other columns
            if col_scrub == "":
                fnames.append(cname)
                header_scrubbed = header_scrubbed + ',' + cname
            else:
                fnames.append(col_scrub)
                header_scrubbed = header_scrubbed + ',' + col_scrub
        # print(cname + '--->' + col_scrub)
        col_scrub = ""
        i = i + 1
    df.columns = fnames

    for_return = (df, header_scrubbed)
    return for_return


if __name__ == '__main__':
    main()
